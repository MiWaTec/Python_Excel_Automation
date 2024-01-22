from openpyxl import load_workbook
from openpyxl.styles import Font

# Open workbook
excel_file_path = "C:/Users/warsi/Desktop/ExcelData\Client_Data_colored.xlsx"
wb = load_workbook(excel_file_path)
ws = wb.active

# Get column index
for cell in ws[1]:
    if cell.internal_value == 'Gender':
        col_index = cell.col_idx
        break

# Color cell value
color_dict = {
    'Female': 'a30000',
    'Male': '0515f0'
}

for cells_in_row in ws.iter_rows(min_row=2):
    cell_value = cells_in_row[col_index-1].internal_value
    if cell_value in list(color_dict):
        cells_in_row[col_index-1].font = Font(color=color_dict[cell_value])

# Save workbook
wb.save("C:/Users/warsi/Desktop/ExcelData\Client_Data.xlsx")