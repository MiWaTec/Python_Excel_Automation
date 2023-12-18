from openpyxl import load_workbook
from openpyxl.styles import Font

# Open workbook
excel_file_path = "C:/Users/Michael/Desktop/ExcelData\Client_Data.xlsx"
wb = load_workbook(excel_file_path)
ws = wb.active

# Define colors
text_color_dict = {
    'Female': 'f0a605',
    'Male': '0515f0'
}

# Get column index
for cell in ws[1]:
    if cell.internal_value == 'Gender':
        col_index = cell.col_idx
        break

# Color text of cells that match the value
for cells_in_row in ws.iter_rows(min_row=2,
                                 min_col=col_index,
                                 max_col=col_index):
    cell_value = cells_in_row[0].internal_value
    if cell_value in list(text_color_dict):
        cells_in_row[0].font = Font(color=text_color_dict[cell_value],
                                    bold=True)

# Color numbers of cells that are in a specific range
for cells_in_row in ws.iter_rows(min_row=2,
                                 min_col=col_index,
                                 max_col=col_index):
    cell_value = cells_in_row[0].internal_value
    if cell_value < 30:
        cells_in_row[0].font = Font(color=text_color_dict[cell_value])
    elif cell_value in range(31, 41):
        cells_in_row[0].font = Font(color=text_color_dict[cell_value])
    else:
        cells_in_row[0].font = Font(color=text_color_dict[cell_value])

# Save workbook
wb.save(excel_file_path)