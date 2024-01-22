from openpyxl import load_workbook
from openpyxl.styles import Font

# Open workbook
excel_file_path = "C:/Users/warsi/Desktop/ExcelData\Client_Data.xlsx"
wb = load_workbook(excel_file_path)
ws = wb.active

# Get column index
for cell in ws[1]:
    if cell.internal_value == 'BMI':
        col_index = cell.col_idx
        break

# Color cell value
color_dict = {
    'underweight': '00bcd1',
    'normal weight': '00a323',
    'overweight': 'fc033d'
}

for cells_in_row in ws.iter_rows(min_row=2):
    cell_value = cells_in_row[col_index-1].internal_value
    if cell_value < 18.5:
        cells_in_row[0].font = Font(color=color_dict['underweight'])
    elif cell_value >= 18.5 and cell_value <= 24.9:
        cells_in_row[0].font = Font(color=color_dict['normal weight'])
    else:
        cells_in_row[0].font = Font(color=color_dict['overweight'])

# Save workbook
wb.save(excel_file_path)