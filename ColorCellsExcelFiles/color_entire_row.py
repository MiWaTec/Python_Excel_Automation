from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN

# Open workbook
excel_file = 'C:/Users/Michael/Desktop/ExcelData\Client_Data1.xlsx'
wb = load_workbook(excel_file)
ws = wb.active

# Get column index 
for cell in ws[1]:
    if cell.internal_value == 'Age':
        col_index = cell.col_idx
        break

# Set border color
border_look = Side(border_style=BORDER_THIN, color='00bbbdbb')
border = Border(
    left=border_look,
    right=border_look,
    top=border_look,
    bottom=border_look
)

# Find all the rows that match a specific value
row_list =[]
for cells_in_row in ws.iter_rows(min_row=2,
                                 min_col=col_index,
                                 max_col=col_index):
        if cells_in_row[0].internal_value <= 30:
            row_list.append(cells_in_row[0].row)

# Color the entire row
for row in row_list:
    for cell in ws[row]:
        cell.fill = PatternFill(patternType='solid', fgColor='0af525')
        cell.border = border

# Save workbook
wb.save(excel_file)
