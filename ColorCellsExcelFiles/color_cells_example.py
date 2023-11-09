import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN

def create_new_excel_file(folder, file_name):
    # Create absolute path for the new excel file
    path = os.path.join(folder, file_name)
    # Color cells in workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Color Cells'
    wb.save(path)
    return path

def color_excel_cells(excel_file_path, color_header, color_1, color_2):
    # Set colors 
    border_look = Side(border_style=BORDER_THIN, color='00bbbdbb')
    border = Border(
        left=border_look,
        right=border_look,
        top=border_look,
        bottom=border_look
    )

    # Open Workbook and color the cells
    wb = load_workbook(excel_file_path)
    ws = wb.active
    number_rows = ws.max_row
    for row in ws.iter_rows(min_row=1, max_row=number_rows):
        if row[1].row == 1:
            color = color_header
        elif row[1].row % 2 == 0:
            color = color_1
        else:
            color = color_2
        for cell in row:
            print(cell)
            print(cell.coordinate)
            print(cell.internal_value)
            print(type(cell.internal_value))
            cell.fill = PatternFill(patternType='solid', fgColor=color)
            cell.border = border
        
    wb.save(excel_file_path)

color_excel_cells('C:/Users/Michael/Desktop/YouTube_Channel/Video_002\Client_List_Test.xlsx',
                  '5a9de8', '9ac5f5', 'ffffff')