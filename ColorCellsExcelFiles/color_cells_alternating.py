from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN

# Open workbook
excel_file = 'C:/Users/Michael/Desktop/ExcelData\Client_Data.xlsx'
wb = load_workbook(excel_file)
ws = wb.active

# Set cell colors
color_header = 'e68e37'
color_1 = 'ffffff'
color_2 = 'ffc48a'

# Set border color
border_look = Side(border_style=BORDER_THIN, color='00bbbdbb')
border = Border(
    left=border_look,
    right=border_look,
    top=border_look,
    bottom=border_look
)

# Color cells in alternating style
number_rows = ws.max_row
for row in ws.iter_rows(min_row=1, max_row=number_rows):
    if row[1].row == 1:
        color = color_header
    elif row[1].row % 2 == 0:
        color = color_1
    else:
        color = color_2
    for cell in row:
        cell.fill = PatternFill(patternType='solid', fgColor=color)
        cell.border = border

# Save workbook
wb.save(excel_file)