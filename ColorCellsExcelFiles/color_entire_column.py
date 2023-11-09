from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN

# Open workbook
excel_file = 'C:/Users/Michael/Desktop/ExcelData\Client_Data.xlsx'
wb = load_workbook(excel_file)
ws = wb.active

# Set border color
border_look = Side(border_style=BORDER_THIN, color='00bbbdbb')
border = Border(
    left=border_look,
    right=border_look,
    top=border_look,
    bottom=border_look
)

# Get column index 
for cell in ws[1]:
    if cell.internal_value == 'Gender':
        col_index = cell.col_idx
        break

# Color cells that match a specific value
for cell_in_row in ws.iter_rows(min_row=1, min_col=col_index, max_col=col_index):
    for cell in cell_in_row:
        cell.fill = PatternFill(patternType='solid', fgColor='ffffff')
        cell.border = border

# Save workbook
wb.save(excel_file)
#'f5ed0a'