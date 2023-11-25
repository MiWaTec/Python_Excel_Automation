from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Open workbook
excel_file = 'C:/Users/Michael/Desktop/ExcelData\Client_Data.xlsx'
wb = load_workbook(excel_file)
ws = wb.active

# Get column index
for cell in ws[1]:
    if cell.internal_value == 'Gender':
        col_index = cell.col_idx
        break

# Color cells that match a specific value
for cells_in_row in ws.iter_rows(min_row=2, 
                                 min_col=col_index,
                                 max_col=col_index):
    if cells_in_row[0].internal_value == 'Female':
        cells_in_row[0].fill = PatternFill(patternType='solid', fgColor='e36464')

# Save workbook
wb.save(excel_file)